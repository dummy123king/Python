import openpyxl
import argparse

# Load workbook
wb = openpyxl.load_workbook('/home/mirafra/gitPull/CPractice/Python/Sample.xlsx')

# Access a worksheet
ws = wb['Sheet1']

# Write the alphabet from A to Z into cells
for letter_ascii in (range(65, 91)):  # ASCII values for A-Z
    letter = chr(letter_ascii)  # Convert ASCII value to letter
    cell_ref = f'{letter}{1}'  # Construct cell reference (e.g., A1, A2, ...)
    
    if ws[cell_ref].value != None:  # Write letter to cell
        print(ws[cell_ref].value)  # Write letter to cell



# Write data to a cell
#ws['A1'] = 'Hello, World!'

# Save workbook
#wb.save('example_modified.xlsx')